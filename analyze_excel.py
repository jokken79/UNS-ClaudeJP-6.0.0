#!/usr/bin/env python3
"""
Analyze Excel employee master file structure
"""
import openpyxl
import sys
from datetime import datetime

excel_path = '/app/BASEDATEJP/【新】社員台帳(UNS)T　2022.04.05～.xlsm'

def analyze_sheet(ws, sheet_name):
    print(f'\n{"=" * 80}')
    print(f'SHEET: {sheet_name}')
    print(f'{"=" * 80}')

    max_row = ws.max_row
    max_col = ws.max_column

    print(f'Dimensions: {max_row} rows x {max_col} columns')

    # Find header row
    header_row = 1
    headers = []
    for col in range(1, min(max_col + 1, 50)):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value:
            headers.append(str(cell_value).strip())
        else:
            headers.append(f'[Col{col}]')

    print(f'\nFirst 30 column headers:')
    for i, h in enumerate(headers[:30], 1):
        print(f'  {i:2d}. {h}')

    # Count non-empty rows (sample first 100)
    non_empty_count = 0
    for row_num in range(header_row + 1, min(header_row + 100, max_row + 1)):
        first_cell = ws.cell(row=row_num, column=1).value
        if first_cell is not None and str(first_cell).strip():
            non_empty_count += 1

    print(f'\nEstimated non-empty rows (sample): {non_empty_count}')

    # Sample first 2 data rows
    print(f'\nSample data (first 2 rows, first 15 columns):')
    for row_idx in range(header_row + 1, min(header_row + 3, max_row + 1)):
        print(f'\n  Row {row_idx}:')
        for col_idx in range(1, min(16, max_col + 1)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            col_name = headers[col_idx - 1] if col_idx - 1 < len(headers) else f'Col{col_idx}'
            if cell_value is not None:
                # Format datetime objects
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime('%Y-%m-%d')
                display_value = str(cell_value)[:50]  # Truncate long values
                print(f'    {col_name}: {display_value}')

def main():
    try:
        print('=' * 80)
        print('EXCEL FILE ANALYSIS')
        print('=' * 80)
        print(f'\nLoading: 【新】社員台帳(UNS)T　2022.04.05～.xlsm')

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        print(f'\nTotal sheets: {len(wb.sheetnames)}')
        print('\nAll sheets:')
        for i, name in enumerate(wb.sheetnames, 1):
            ws = wb[name]
            print(f'  {i:2d}. {name}: {ws.max_row} rows x {ws.max_column} cols')

        # Analyze main sheets with data
        main_sheets = ['DBGenzaiX', '派遣社員', 'DBUkeoiX', '請負社員', 'DBStaffX', 'スタッフ']

        for sheet_name in main_sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                analyze_sheet(ws, sheet_name)

        wb.close()
        print('\n' + '=' * 80)
        print('ANALYSIS COMPLETE')
        print('=' * 80)

    except Exception as e:
        import traceback
        print(f'\nERROR: {str(e)}', file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()
