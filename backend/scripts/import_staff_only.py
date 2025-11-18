#!/usr/bin/env python3
"""
Import only Staff (スタッフ) from DBStaffX sheet
"""
import openpyxl
import sys
from datetime import datetime, timedelta
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, '/app')

from app.models.models import Staff
from app.core.config import settings

# Excel constants (from 1900-01-01 offset)
EXCEL_EPOCH = datetime(1900, 1, 1)

def excel_date_to_datetime(excel_date):
    """Convert Excel date number to datetime"""
    if not excel_date or excel_date == 0:
        return None
    try:
        if isinstance(excel_date, (int, float)):
            # Excel dates start from Jan 1, 1900 = 1
            # But Excel has a leap year bug for Feb 29, 1900
            if excel_date < 60:
                adjusted_date = EXCEL_EPOCH + timedelta(days=excel_date - 1)
            else:
                adjusted_date = EXCEL_EPOCH + timedelta(days=excel_date - 2)
            return adjusted_date.date()
    except:
        pass

    # Try parsing if it's already a datetime
    if isinstance(excel_date, datetime):
        return excel_date.date()

    return None

def load_sheet_data(excel_path, sheet_name):
    """Load data from Excel sheet"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    data = []
    header_row = None

    # Find header row
    for row in range(1, min(10, ws.max_row + 1)):
        values = [cell.value for cell in ws[row]]
        if any(v is not None for v in values):
            header_row = row
            break

    if not header_row:
        wb.close()
        return []

    # Get headers
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]

    # Get data rows
    for row in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        has_data = False

        for col, header in enumerate(headers, 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                has_data = True
            row_data[header] = value

        if has_data:
            data.append(row_data)

    wb.close()
    return data

def import_staff(excel_path):
    """Import Staff from DBStaffX sheet"""

    engine = create_engine(settings.DATABASE_URL)
    SessionLocal = sessionmaker(bind=engine)
    session = SessionLocal()

    print("\n" + "=" * 80)
    print("IMPORTING STAFF (スタッフ)")
    print("=" * 80)

    try:
        # Load staff data from DBStaffX
        print("\n[*] Loading staff from DBStaffX...")
        staff_data = load_sheet_data(excel_path, 'DBStaffX')
        print(f"  [OK] Loaded {len(staff_data)} staff records")

        print("\n[*] Importing staff...")
        staff_id_counter = 1  # Start from 1 for staff_id (unique field for Staff model)
        imported_count = 0
        error_count = 0

        for idx, row in enumerate(staff_data, 1):
            try:
                name = row.get('氏名')
                if not name or str(name).strip() == '':
                    continue

                staff = Staff(
                    staff_id=staff_id_counter,  # Staff uses staff_id, not hakenmoto_id
                    full_name_kanji=str(name).strip(),
                    full_name_kana=str(row.get('カナ', '')).strip() or None,
                    gender=str(row.get('性別', '')).strip() or None,
                    nationality=str(row.get('国籍', '')).strip() or None,
                    date_of_birth=excel_date_to_datetime(row.get('生年月日'))
                )

                session.add(staff)
                staff_id_counter += 1
                imported_count += 1

            except Exception as e:
                session.rollback()
                error_count += 1
                if error_count <= 5:
                    print(f"  [!] Error at row {idx}: {str(e)[:80]}")

        session.commit()
        print(f"  [OK] Staff import complete: {imported_count} records")

        # Verify import
        total_staff = session.query(Staff).count()

        print("\n" + "=" * 80)
        print("IMPORT RESULTS")
        print("=" * 80)
        print(f"Staff imported: {imported_count}")
        print(f"Total in database: {total_staff}")
        print(f"Errors: {error_count}")
        print("=" * 80 + "\n")

        session.close()
        return imported_count > 0

    except Exception as e:
        print(f"\n[X] Critical error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    excel_file = "/app/BASEDATEJP/【新】社員台帳(UNS)T　2022.04.05～.xlsm"
    success = import_staff(excel_file)
    sys.exit(0 if success else 1)
