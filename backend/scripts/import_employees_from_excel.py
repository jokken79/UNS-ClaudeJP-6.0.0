"""
Import employees from Excel master file to PostgreSQL
Imports 1,050 employees + 100 contract workers + 16 staff

File: BASEDATEJP/【新】社員台帳(UNS)T　2022.04.05～.xlsm
Sheets: DBGenzaiX (employees), DBUkeoiX (contract_workers), DBStaffX (staff)
"""
import sys
import os
from pathlib import Path
from datetime import datetime, date
import time
import openpyxl
from sqlalchemy.orm import Session
from typing import Optional, Dict, Any

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.core.database import SessionLocal, engine
from app.models.models import Employee, ContractWorker, Staff, Factory, Apartment


def convert_excel_date(excel_date) -> Optional[date]:
    """Convert Excel serial date to Python date object"""
    if excel_date is None or excel_date == 0 or excel_date == "0":
        return None

    # If already a datetime object
    if isinstance(excel_date, datetime):
        return excel_date.date()

    # If already a date object
    if isinstance(excel_date, date):
        return excel_date

    # If string, try to parse
    if isinstance(excel_date, str):
        try:
            # Try YYYY-MM-DD format
            return datetime.strptime(excel_date, "%Y-%m-%d").date()
        except:
            try:
                # Try YYYY/MM/DD format
                return datetime.strptime(excel_date, "%Y/%m/%d").date()
            except:
                return None

    # If numeric (Excel serial date)
    try:
        excel_epoch = datetime(1899, 12, 30)
        return (excel_epoch + timedelta(days=int(excel_date))).date()
    except:
        return None


def clean_value(value, convert_zero_to_null=True):
    """Clean Excel values - convert 0/"0" to NULL, strip whitespace"""
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()
        if value == "" or (convert_zero_to_null and value == "0"):
            return None

    if convert_zero_to_null and value == 0:
        return None

    return value


def clean_integer(value) -> Optional[int]:
    """Convert to integer, handle NULL/0"""
    cleaned = clean_value(value)
    if cleaned is None:
        return None
    try:
        return int(cleaned)
    except:
        return None


def get_factory_id_by_name(db: Session, factory_name: str) -> Optional[str]:
    """Lookup factory_id by company name"""
    if not factory_name:
        return None

    # Search in factories table
    factory = db.query(Factory).filter(
        Factory.company_name.ilike(f"%{factory_name}%")
    ).first()

    return factory.factory_id if factory else None


def import_employees(db: Session, wb: openpyxl.Workbook) -> Dict[str, Any]:
    """Import employees from DBGenzaiX sheet"""
    print("\n=== Importing Employees (DBGenzaiX) ===")

    ws = wb['DBGenzaiX']
    total_rows = ws.max_row - 1  # Exclude header
    imported = 0
    errors = []
    batch = []
    batch_size = 100

    for row_num in range(2, ws.max_row + 1):
        try:
            # Read row data
            row = [ws.cell(row_num, col).value for col in range(1, ws.max_column + 1)]

            # Extract fields
            status_text = clean_value(row[0])  # 現況
            hakenmoto_id = clean_integer(row[1])  # 社員番号
            factory_name = clean_value(row[3])  # 派遣先
            full_name_roman = clean_value(row[7])  # 氏名
            full_name_kana = clean_value(row[8])  # カナ
            gender = clean_value(row[9])  # 性別
            nationality = clean_value(row[10])  # 国籍
            date_of_birth_raw = row[11]  # 生年月日
            jikyu = clean_integer(row[13])  # 時給
            jikyu_revision_text = row[14]  # 時給改定
            hourly_rate_charged = clean_integer(row[15])  # 請求単価
            billing_revision_text = row[16]  # 請求改定
            profit_difference = clean_integer(row[17])  # 差額利益
            standard_compensation = clean_integer(row[18])  # 標準報酬
            visa_type = clean_value(row[22])  # ビザ種類
            zairyu_expire_raw = row[23]  # 期限
            phone = clean_value(row[25])  # 電
            address = clean_value(row[26])  # 住所
            current_hire_date_raw = row[27]  # 現入社
            social_insurance_date_raw = row[32] if len(row) > 32 else None  # 社保加入日
            entry_request_date_raw = row[33] if len(row) > 33 else None  # 入社依頼

            # Skip if no employee ID
            if not hakenmoto_id or hakenmoto_id == 0:
                continue

            # Check if employee already exists
            existing = db.query(Employee).filter(
                Employee.hakenmoto_id == hakenmoto_id,
                Employee.is_deleted == False
            ).first()

            if existing:
                print(f"  Row {row_num}: Employee {hakenmoto_id} already exists - skipping")
                continue

            # Determine status
            is_active = True
            termination_date = None
            if status_text and "退社" in status_text:
                is_active = False
                # Try to extract termination date from status text if available

            # Convert dates
            date_of_birth = convert_excel_date(date_of_birth_raw)
            current_hire_date = convert_excel_date(current_hire_date_raw)
            zairyu_expire_date = convert_excel_date(zairyu_expire_raw)
            social_insurance_date = convert_excel_date(social_insurance_date_raw)
            entry_request_date = convert_excel_date(entry_request_date_raw)

            # Parse jikyu_revision_date from text (e.g., "2020/11/21 1200→1250")
            jikyu_revision_date = None
            if isinstance(jikyu_revision_text, str) and "/" in jikyu_revision_text:
                try:
                    date_part = jikyu_revision_text.split()[0]
                    jikyu_revision_date = datetime.strptime(date_part, "%Y/%m/%d").date()
                except:
                    pass
            elif jikyu_revision_text:
                jikyu_revision_date = convert_excel_date(jikyu_revision_text)

            # Parse billing_revision_date
            billing_revision_date = None
            if isinstance(billing_revision_text, str) and "/" in billing_revision_text:
                try:
                    date_part = billing_revision_text.split()[0]
                    billing_revision_date = datetime.strptime(date_part, "%Y/%m/%d").date()
                except:
                    pass
            elif billing_revision_text:
                billing_revision_date = convert_excel_date(billing_revision_text)

            # Lookup factory_id
            factory_id = get_factory_id_by_name(db, factory_name)

            # Determine full_name_kanji (use kana if roman name looks like kanji)
            full_name_kanji = full_name_roman  # Most names are actually in Roman/Kanji mix

            # Create employee object
            employee = Employee(
                hakenmoto_id=hakenmoto_id,
                factory_id=factory_id,
                company_name=factory_name,
                full_name_kanji=full_name_kanji,
                full_name_kana=full_name_kana,
                gender=gender,
                nationality=nationality,
                date_of_birth=date_of_birth,
                jikyu=jikyu,
                jikyu_revision_date=jikyu_revision_date,
                hourly_rate_charged=hourly_rate_charged,
                billing_revision_date=billing_revision_date,
                profit_difference=profit_difference,
                standard_compensation=standard_compensation,
                visa_type=visa_type,
                zairyu_expire_date=zairyu_expire_date,
                phone=phone,
                address=address,
                current_hire_date=current_hire_date,
                social_insurance_date=social_insurance_date,
                entry_request_date=entry_request_date,
                is_active=is_active,
                termination_date=termination_date,
                hire_date=current_hire_date,  # Use current_hire_date as hire_date
            )

            batch.append(employee)
            imported += 1

            # Print progress every 50 records
            if imported % 50 == 0:
                print(f"  Processed {imported}/{total_rows} employees...")

            # Batch insert
            if len(batch) >= batch_size:
                db.add_all(batch)
                db.commit()
                batch = []

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
            print(f"  ⚠ Error at row {row_num}: {str(e)}")

    # Insert remaining batch
    if batch:
        db.add_all(batch)
        db.commit()

    return {
        "total_rows": total_rows,
        "imported": imported,
        "errors": errors
    }


def import_contract_workers(db: Session, wb: openpyxl.Workbook) -> Dict[str, Any]:
    """Import contract workers from DBUkeoiX sheet"""
    print("\n=== Importing Contract Workers (DBUkeoiX) ===")

    ws = wb['DBUkeoiX']
    total_rows = ws.max_row - 1  # Exclude header
    imported = 0
    errors = []
    batch = []
    batch_size = 100

    for row_num in range(2, ws.max_row + 1):
        try:
            # Read row data
            row = [ws.cell(row_num, col).value for col in range(1, ws.max_column + 1)]

            # Extract fields
            status_text = clean_value(row[0])  # 現況
            hakenmoto_id = clean_integer(row[1])  # 社員番号
            full_name_kanji = clean_value(row[3])  # 氏名
            full_name_kana = clean_value(row[4])  # カナ
            gender = clean_value(row[5])  # 性別
            nationality = clean_value(row[6])  # 国籍
            date_of_birth_raw = row[7]  # 生年月日
            jikyu = clean_integer(row[9])  # 時給
            jikyu_revision_raw = row[10]  # 時給改定
            visa_type = clean_value(row[18]) if len(row) > 18 else None  # ビザ種類
            zairyu_expire_raw = row[19] if len(row) > 19 else None  # 期限

            # Skip if no employee ID
            if not hakenmoto_id or hakenmoto_id == 0:
                continue

            # Check if already exists
            existing = db.query(ContractWorker).filter(
                ContractWorker.hakenmoto_id == hakenmoto_id,
                ContractWorker.is_deleted == False
            ).first()

            if existing:
                print(f"  Row {row_num}: Contract worker {hakenmoto_id} already exists - skipping")
                continue

            # Determine status
            is_active = True
            if status_text and "退社" in status_text:
                is_active = False

            # Convert dates
            date_of_birth = convert_excel_date(date_of_birth_raw)
            jikyu_revision_date = convert_excel_date(jikyu_revision_raw)
            zairyu_expire_date = convert_excel_date(zairyu_expire_raw)

            # Create contract worker object
            contract_worker = ContractWorker(
                hakenmoto_id=hakenmoto_id,
                full_name_kanji=full_name_kanji,
                full_name_kana=full_name_kana,
                gender=gender,
                nationality=nationality,
                date_of_birth=date_of_birth,
                jikyu=jikyu,
                jikyu_revision_date=jikyu_revision_date,
                visa_type=visa_type,
                zairyu_expire_date=zairyu_expire_date,
                is_active=is_active,
            )

            batch.append(contract_worker)
            imported += 1

            # Batch insert
            if len(batch) >= batch_size:
                db.add_all(batch)
                db.commit()
                batch = []

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
            print(f"  ⚠ Error at row {row_num}: {str(e)}")

    # Insert remaining batch
    if batch:
        db.add_all(batch)
        db.commit()

    return {
        "total_rows": total_rows,
        "imported": imported,
        "errors": errors
    }


def import_staff(db: Session, wb: openpyxl.Workbook) -> Dict[str, Any]:
    """Import staff from DBStaffX sheet"""
    print("\n=== Importing Staff (DBStaffX) ===")

    ws = wb['DBStaffX']
    total_rows = ws.max_row - 1  # Exclude header
    imported = 0
    errors = []
    batch = []

    for row_num in range(2, ws.max_row + 1):
        try:
            # Read row data
            row = [ws.cell(row_num, col).value for col in range(1, ws.max_column + 1)]

            # Extract fields
            staff_id = clean_integer(row[1])  # 社員番号
            full_name_kanji = clean_value(row[3])  # 氏名
            full_name_kana = clean_value(row[4])  # カナ
            gender = clean_value(row[5])  # 性別
            nationality = clean_value(row[6])  # 国籍
            date_of_birth_raw = row[7]  # 生年月日
            visa_type = clean_value(row[9]) if len(row) > 9 else None  # ビザ種類
            phone = clean_value(row[12]) if len(row) > 12 else None  # 電
            address = clean_value(row[13]) if len(row) > 13 else None  # 住所

            # Skip if no staff ID
            if not staff_id or staff_id == 0:
                continue

            # Check if already exists
            existing = db.query(Staff).filter(
                Staff.staff_id == staff_id,
                Staff.is_deleted == False
            ).first()

            if existing:
                print(f"  Row {row_num}: Staff {staff_id} already exists - skipping")
                continue

            # Convert dates
            date_of_birth = convert_excel_date(date_of_birth_raw)

            # Create staff object
            staff = Staff(
                staff_id=staff_id,
                full_name_kanji=full_name_kanji,
                full_name_kana=full_name_kana,
                gender=gender,
                nationality=nationality,
                date_of_birth=date_of_birth,
                phone=phone,
                address=address,
                is_active=True,
            )

            batch.append(staff)
            imported += 1

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
            print(f"  ⚠ Error at row {row_num}: {str(e)}")

    # Insert all staff
    if batch:
        db.add_all(batch)
        db.commit()

    return {
        "total_rows": total_rows,
        "imported": imported,
        "errors": errors
    }


def verify_import(db: Session):
    """Verify import counts"""
    print("\n=== Verification ===")

    employee_count = db.query(Employee).filter(Employee.is_deleted == False).count()
    contract_worker_count = db.query(ContractWorker).filter(ContractWorker.is_deleted == False).count()
    staff_count = db.query(Staff).filter(Staff.is_deleted == False).count()

    print(f"  Employees: {employee_count}")
    print(f"  Contract Workers: {contract_worker_count}")
    print(f"  Staff: {staff_count}")
    print(f"  TOTAL: {employee_count + contract_worker_count + staff_count}")

    return {
        "employees": employee_count,
        "contract_workers": contract_worker_count,
        "staff": staff_count,
        "total": employee_count + contract_worker_count + staff_count
    }


def main():
    """Main import function"""
    start_time = time.time()

    # Excel file path (inside Docker container)
    excel_path = Path("/app/BASEDATEJP/【新】社員台帳(UNS)T　2022.04.05～.xlsm")

    print(f"Reading Excel file: {excel_path}")

    if not excel_path.exists():
        print(f"❌ ERROR: Excel file not found at {excel_path}")
        return

    # Load Excel workbook
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    print(f"Available sheets: {wb.sheetnames}")

    # Create database session
    db = SessionLocal()

    try:
        # Import employees
        employee_result = import_employees(db, wb)
        print(f"  ✓ Imported {employee_result['imported']}/{employee_result['total_rows']} employees")
        if employee_result['errors']:
            print(f"  ⚠ {len(employee_result['errors'])} errors")

        # Import contract workers
        contract_result = import_contract_workers(db, wb)
        print(f"  ✓ Imported {contract_result['imported']}/{contract_result['total_rows']} contract workers")
        if contract_result['errors']:
            print(f"  ⚠ {len(contract_result['errors'])} errors")

        # Import staff
        staff_result = import_staff(db, wb)
        print(f"  ✓ Imported {staff_result['imported']}/{staff_result['total_rows']} staff")
        if staff_result['errors']:
            print(f"  ⚠ {len(staff_result['errors'])} errors")

        # Verify import
        verification = verify_import(db)

        # Calculate execution time
        elapsed_time = time.time() - start_time

        print(f"\n{'='*60}")
        print(f"✓ Import completed in {elapsed_time:.2f} seconds")
        print(f"{'='*60}")

        # Print error summary
        all_errors = employee_result['errors'] + contract_result['errors'] + staff_result['errors']
        if all_errors:
            print(f"\n⚠ Total errors: {len(all_errors)}")
            print("First 10 errors:")
            for error in all_errors[:10]:
                print(f"  - {error}")

    except Exception as e:
        print(f"\n❌ CRITICAL ERROR: {str(e)}")
        db.rollback()
        raise

    finally:
        db.close()
        wb.close()


if __name__ == "__main__":
    from datetime import timedelta
    main()
