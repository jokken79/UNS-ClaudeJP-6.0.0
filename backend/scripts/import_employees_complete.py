#!/usr/bin/env python3
"""
Comprehensive import of Employees, Contract Workers, and Staff from Excel
with proper hakenmoto_id assignment
"""
import openpyxl
import sys
from datetime import datetime, timedelta
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, '/app')

from app.models.models import Employee, ContractWorker, Staff
from app.core.config import settings

# Excel constants (from 1900-01-01 offset)
EXCEL_EPOCH = datetime(1900, 1, 1)

def excel_date_to_datetime(excel_date):
    """Convert Excel date number to datetime"""
    if not excel_date or excel_date == 0:
        return None
    try:
        if isinstance(excel_date, (int, float)):
            # Excel dates start from Jan 1, 1900 = 1
            # But Excel has a leap year bug for Feb 29, 1900
            if excel_date < 60:
                adjusted_date = EXCEL_EPOCH + timedelta(days=excel_date - 1)
            else:
                adjusted_date = EXCEL_EPOCH + timedelta(days=excel_date - 2)
            return adjusted_date.date()
    except:
        pass

    # Try parsing if it's already a datetime
    if isinstance(excel_date, datetime):
        return excel_date.date()

    return None

def load_sheet_data(excel_path, sheet_name):
    """Load data from Excel sheet"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    data = []
    header_row = None

    # Find header row
    for row in range(1, min(10, ws.max_row + 1)):
        values = [cell.value for cell in ws[row]]
        if any(v is not None for v in values):
            header_row = row
            break

    if not header_row:
        wb.close()
        return []

    # Get headers
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]

    # Get data rows
    for row in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        has_data = False

        for col, header in enumerate(headers, 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                has_data = True
            row_data[header] = value

        if has_data:
            data.append(row_data)

    wb.close()
    return data

def import_employees(excel_path):
    """Import all employee types from Excel"""

    engine = create_engine(settings.DATABASE_URL)
    SessionLocal = sessionmaker(bind=engine)
    session = SessionLocal()

    print("\n" + "=" * 80)
    print("IMPORTING EMPLOYEES, CONTRACT WORKERS, AND STAFF")
    print("=" * 80)

    try:
        hakenmoto_counter = 1  # Start from 1 for unique hakenmoto_id
        stats = {
            'employees': 0,
            'contract_workers': 0,
            'staff': 0,
            'errors': 0
        }

        # 1. Import Employees (DBGenzaiX)
        print("\n[*] Loading employees from DBGenzaiX...")
        employee_data = load_sheet_data(excel_path, 'DBGenzaiX')
        print(f"  [OK] Loaded {len(employee_data)} employee records")

        print("\n[*] Importing employees...")
        for idx, row in enumerate(employee_data, 1):
            try:
                name = row.get('氏名')
                if not name or str(name).strip() == '':
                    continue

                emp = Employee(
                    hakenmoto_id=hakenmoto_counter,
                    hakensaki_shain_id=str(row.get('社員№', '')).strip() or None,
                    full_name_kanji=str(name).strip(),
                    full_name_kana=str(row.get('カナ', '')).strip() or None,
                    gender=str(row.get('性別', '')).strip() or None,
                    nationality=str(row.get('国籍', '')).strip() or None,
                    date_of_birth=excel_date_to_datetime(row.get('生年月日')),
                    jikyu=row.get('時給'),
                    is_active=True,
                    is_corporate_housing=False
                )

                session.add(emp)
                hakenmoto_counter += 1
                stats['employees'] += 1

                if stats['employees'] % 100 == 0:
                    session.commit()
                    print(f"  [OK] Imported {stats['employees']} employees")

            except Exception as e:
                session.rollback()
                stats['errors'] += 1
                if stats['errors'] <= 5:
                    print(f"  [!] Error at row {idx}: {str(e)[:80]}")

        session.commit()
        print(f"  [OK] Employees import complete: {stats['employees']} records")

        # 2. Import Contract Workers (DBUkeoiX)
        print("\n[*] Loading contract workers from DBUkeoiX...")
        contract_data = load_sheet_data(excel_path, 'DBUkeoiX')
        print(f"  [OK] Loaded {len(contract_data)} contract worker records")

        print("\n[*] Importing contract workers...")
        for idx, row in enumerate(contract_data, 1):
            try:
                name = row.get('氏名')
                if not name or str(name).strip() == '':
                    continue

                cw = ContractWorker(
                    hakenmoto_id=hakenmoto_counter,
                    hakensaki_shain_id=str(row.get('社員№', '')).strip() or None,
                    full_name_kanji=str(name).strip(),
                    full_name_kana=str(row.get('カナ', '')).strip() or None,
                    gender=str(row.get('性別', '')).strip() or None,
                    nationality=str(row.get('国籍', '')).strip() or None,
                    date_of_birth=excel_date_to_datetime(row.get('生年月日')),
                    jikyu=row.get('時給'),
                    is_active=True,
                    is_corporate_housing=False
                )

                session.add(cw)
                hakenmoto_counter += 1
                stats['contract_workers'] += 1

                if stats['contract_workers'] % 50 == 0:
                    session.commit()
                    print(f"  [OK] Imported {stats['contract_workers']} contract workers")

            except Exception as e:
                session.rollback()
                stats['errors'] += 1
                if stats['errors'] <= 5:
                    print(f"  [!] Error at row {idx}: {str(e)[:80]}")

        session.commit()
        print(f"  [OK] Contract workers import complete: {stats['contract_workers']} records")

        # 3. Import Staff (DBStaffX)
        print("\n[*] Loading staff from DBStaffX...")
        staff_data = load_sheet_data(excel_path, 'DBStaffX')
        print(f"  [OK] Loaded {len(staff_data)} staff records")

        print("\n[*] Importing staff...")
        staff_id_counter = 1  # Start from 1 for staff_id (unique field for Staff model)
        for idx, row in enumerate(staff_data, 1):
            try:
                name = row.get('氏名')
                if not name or str(name).strip() == '':
                    continue

                staff = Staff(
                    staff_id=staff_id_counter,  # Staff uses staff_id, not hakenmoto_id
                    full_name_kanji=str(name).strip(),
                    full_name_kana=str(row.get('カナ', '')).strip() or None,
                    gender=str(row.get('性別', '')).strip() or None,
                    nationality=str(row.get('国籍', '')).strip() or None,
                    date_of_birth=excel_date_to_datetime(row.get('生年月日'))
                )

                session.add(staff)
                staff_id_counter += 1
                stats['staff'] += 1

            except Exception as e:
                session.rollback()
                stats['errors'] += 1
                if stats['errors'] <= 5:
                    print(f"  [!] Error at row {idx}: {str(e)[:80]}")

        session.commit()
        print(f"  [OK] Staff import complete: {stats['staff']} records")

        # 4. Verify imports
        total_employees = session.query(Employee).count()
        total_contract_workers = session.query(ContractWorker).count()
        total_staff = session.query(Staff).count()

        print("\n" + "=" * 80)
        print("IMPORT RESULTS")
        print("=" * 80)
        print(f"Employees:        {total_employees} (imported: {stats['employees']})")
        print(f"Contract workers: {total_contract_workers} (imported: {stats['contract_workers']})")
        print(f"Staff:            {total_staff} (imported: {stats['staff']})")
        print(f"Total:            {total_employees + total_contract_workers + total_staff}")
        print(f"Errors:           {stats['errors']}")
        print("=" * 80 + "\n")

        session.close()
        return stats

    except Exception as e:
        print(f"\n[X] Critical error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    excel_file = "/app/BASEDATEJP/【新】社員台帳(UNS)T　2022.04.05～.xlsm"
    result = import_employees(excel_file)

    if result:
        total = result['employees'] + result['contract_workers'] + result['staff']
        if total > 0:
            print("[OK] Import successful!")
            sys.exit(0)

    print("[X] Import failed")
    sys.exit(1)
