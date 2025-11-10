"""
Script para comparar archivos Excel y mostrar todas las hojas (incluyendo ocultas)
"""
import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

def analyze_excel(file_path, label):
    """Analiza un archivo Excel y muestra todas sus hojas"""
    print("\n" + "=" * 80)
    print(f"📊 ANÁLISIS DE: {label}")
    print("=" * 80)
    print(f"Archivo: {file_path}")

    # Abrir con openpyxl para detectar hojas ocultas
    wb = openpyxl.load_workbook(file_path, read_only=False, keep_vba=True)

    print(f"\n📑 Total de hojas: {len(wb.sheetnames)}")

    sheet_info = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Determinar visibilidad
        visibility = "visible"
        if hasattr(sheet, 'sheet_state'):
            if sheet.sheet_state == 'hidden':
                visibility = "🔒 OCULTA"
            elif sheet.sheet_state == 'veryHidden':
                visibility = "🔒🔒 MUY OCULTA"

        # Contar filas y columnas con datos
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            rows = len(df)
            cols = len(df.columns)
        except Exception as e:
            rows = "ERROR"
            cols = "ERROR"

        sheet_info.append({
            'name': sheet_name,
            'visibility': visibility,
            'rows': rows,
            'cols': cols
        })

    # Mostrar información de cada hoja
    print("\n" + "-" * 80)
    print(f"{'Hoja':<40} {'Estado':<15} {'Filas':<10} {'Columnas':<10}")
    print("-" * 80)

    for info in sheet_info:
        print(f"{info['name']:<40} {info['visibility']:<15} {str(info['rows']):<10} {str(info['cols']):<10}")

    # Detalles de hojas relevantes
    relevant_sheets = ['派遣社員', '請負社員', 'スタッフ', '派遣', '請負', 'スタッフ',
                       '社員', 'データ', 'data', 'Data', 'DATA']

    print("\n" + "=" * 80)
    print("📋 DETALLES DE HOJAS CLAVE")
    print("=" * 80)

    for sheet_name in wb.sheetnames:
        # Mostrar detalles de hojas relevantes o todas si el nombre contiene kanji
        if any(keyword in sheet_name for keyword in relevant_sheets) or any('\u4e00' <= c <= '\u9fff' for c in sheet_name):
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                print(f"\n📄 Hoja: '{sheet_name}'")
                print(f"   Columnas ({len(df.columns)}): {', '.join(df.columns[:10].tolist())}")
                if len(df.columns) > 10:
                    print(f"   ... y {len(df.columns) - 10} columnas más")
                print(f"   Primeras filas: {len(df)} de {pd.read_excel(file_path, sheet_name=sheet_name).shape[0]}")
            except Exception as e:
                print(f"\n📄 Hoja: '{sheet_name}' - ⚠️ Error leyendo: {str(e)[:100]}")

    wb.close()
    return sheet_info


def main():
    print("\n" + "🔍" * 40)
    print("COMPARACIÓN DE ARCHIVOS EXCEL - EMPLEADOS UNS")
    print("🔍" * 40)

    # Archivo actual
    current_file = "/app/config/employee_master.xlsm"

    # Archivo nuevo
    new_file = "/app/config/employee_master_NEW.xlsm"

    try:
        current_info = analyze_excel(current_file, "ARCHIVO ACTUAL (employee_master.xlsm)")
    except Exception as e:
        print(f"\n❌ Error analizando archivo actual: {e}")
        current_info = []

    try:
        new_info = analyze_excel(new_file, "ARCHIVO NUEVO (【新】社員台帳(UNS)T.xlsm)")
    except Exception as e:
        print(f"\n❌ Error analizando archivo nuevo: {e}")
        new_info = []

    # Comparación
    print("\n" + "=" * 80)
    print("📊 COMPARACIÓN RÁPIDA")
    print("=" * 80)

    current_sheets = {s['name'] for s in current_info}
    new_sheets = {s['name'] for s in new_info}

    only_in_current = current_sheets - new_sheets
    only_in_new = new_sheets - current_sheets
    common_sheets = current_sheets & new_sheets

    print(f"\n✅ Hojas en común: {len(common_sheets)}")
    for sheet in sorted(common_sheets):
        print(f"   - {sheet}")

    if only_in_current:
        print(f"\n⚠️  Solo en ACTUAL ({len(only_in_current)}):")
        for sheet in sorted(only_in_current):
            print(f"   - {sheet}")

    if only_in_new:
        print(f"\n✨ Solo en NUEVO ({len(only_in_new)}):")
        for sheet in sorted(only_in_new):
            print(f"   - {sheet}")


if __name__ == "__main__":
    main()
