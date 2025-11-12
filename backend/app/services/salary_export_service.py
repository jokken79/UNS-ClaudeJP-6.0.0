"""
SalaryExportService for UNS-ClaudeJP 5.4.1
Excel and PDF export functionality for salary reports
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.models.models import SalaryCalculation

logger = logging.getLogger(__name__)


class SalaryExportService:
    """Export salary data to Excel and PDF formats"""
    
    @staticmethod
    def export_to_excel(
        salaries: List[SalaryCalculation],
        period_start: str,
        period_end: str
    ) -> BytesIO:
        """
        Exporta salarios a Excel con múltiples sheets
        
        Args:
            salaries: List of SalaryCalculation objects
            period_start: Start date (YYYY-MM-DD)
            period_end: End date (YYYY-MM-DD)
            
        Returns:
            BytesIO: Excel file content
        """
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        try:
            logger.info(f"Exporting {len(salaries)} salaries to Excel")
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # Define styles
            header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
            header_font = Font(bold=True, color="ffffff", size=11)
            
            summary_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
            summary_font = Font(bold=True, size=10)
            
            total_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
            total_font = Font(bold=True, color="ffffff", size=11)
            
            center_alignment = Alignment(horizontal="center", vertical="center")
            right_alignment = Alignment(horizontal="right", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Sheet 1: Resumen Ejecutivo
            ws1 = wb.create_sheet("Resumen")
            
            ws1['A1'] = "RESUMEN EJECUTIVO DE SALARIOS"
            ws1['A1'].font = Font(bold=True, size=14)
            ws1['A1'].alignment = center_alignment
            ws1.merge_cells('A1:F1')
            
            ws1['A2'] = f"Período: {period_start} a {period_end}"
            ws1['A2'].font = Font(italic=True)
            ws1.merge_cells('A2:F2')
            
            ws1['A4'] = "Estadísticas"
            ws1['A4'].font = Font(bold=True, size=12, color="1e3a8a")
            
            ws1['A5'] = "Total de Salarios:"
            ws1['B5'] = len(salaries)
            ws1['A6'] = "Salario Bruto Total:"
            ws1['B6'] = sum(s.gross_salary or 0 for s in salaries)
            ws1['A7'] = "Deducciones Totales:"
            ws1['B7'] = sum((s.apartment_deduction or 0) + (s.other_deductions or 0) for s in salaries)
            ws1['A8'] = "Salario Neto Total:"
            ws1['B8'] = sum(s.net_salary or 0 for s in salaries)
            ws1['A9'] = "Promedio Salario Neto:"
            ws1['B9'] = sum(s.net_salary or 0 for s in salaries) / len(salaries) if salaries else 0
            ws1['A10'] = "Utilidad Empresa Total:"
            ws1['B10'] = sum(s.company_profit or 0 for s in salaries)
            
            # Format currency columns
            for row in range(6, 11):
                ws1[f'B{row}'].number_format = '¥#,##0'
            
            # Sheet 2: Detalle por Empleado
            ws2 = wb.create_sheet("Detalle")
            
            headers = ["ID Empleado", "Nombre", "Período", "Horas Regulares", "Horas Extra",
                      "Salario Bruto", "Apartamento", "Otras Deducciones", "Salario Neto", "Pagado"]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            for row_num, salary in enumerate(salaries, 2):
                ws2.cell(row=row_num, column=1).value = salary.employee_id
                ws2.cell(row=row_num, column=2).value = f"{salary.year}-{salary.month:02d}"
                ws2.cell(row=row_num, column=3).value = salary.total_regular_hours
                ws2.cell(row=row_num, column=4).value = salary.total_overtime_hours
                ws2.cell(row=row_num, column=5).value = salary.gross_salary
                ws2.cell(row=row_num, column=6).value = salary.apartment_deduction
                ws2.cell(row=row_num, column=7).value = salary.other_deductions
                ws2.cell(row=row_num, column=8).value = salary.net_salary
                ws2.cell(row=row_num, column=9).value = "Sí" if salary.is_paid else "No"
                
                # Format currency
                for col in [5, 6, 7, 8]:
                    ws2.cell(row=row_num, column=col).number_format = '¥#,##0'
            
            # Adjust column widths
            ws2.column_dimensions['A'].width = 12
            ws2.column_dimensions['B'].width = 15
            ws2.column_dimensions['C'].width = 12
            ws2.column_dimensions['D'].width = 14
            ws2.column_dimensions['E'].width = 12
            ws2.column_dimensions['F'].width = 14
            ws2.column_dimensions['G'].width = 14
            ws2.column_dimensions['H'].width = 14
            ws2.column_dimensions['I'].width = 14
            ws2.column_dimensions['J'].width = 10
            
            # Sheet 3: Análisis Fiscal
            ws3 = wb.create_sheet("Análisis Fiscal")
            
            ws3['A1'] = "ANÁLISIS DE DEDUCCIONES FISCALES"
            ws3['A1'].font = Font(bold=True, size=12, color="991b1b")
            ws3.merge_cells('A1:D1')
            
            deduction_headers = ["Tipo de Deducción", "Cantidad", "Monto Total", "% del Bruto"]
            for col_num, header in enumerate(deduction_headers, 1):
                cell = ws3.cell(row=3, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            total_gross = sum(s.gross_salary or 0 for s in salaries)
            total_apartment = sum(s.apartment_deduction or 0 for s in salaries)
            total_other = sum(s.other_deductions or 0 for s in salaries)
            
            deductions = [
                ("Apartamento", len([s for s in salaries if s.apartment_deduction]), total_apartment),
                ("Otras Deducciones", len([s for s in salaries if s.other_deductions]), total_other),
            ]
            
            for row_num, (name, count, amount) in enumerate(deductions, 4):
                ws3.cell(row=row_num, column=1).value = name
                ws3.cell(row=row_num, column=2).value = count
                ws3.cell(row=row_num, column=3).value = amount
                ws3.cell(row=row_num, column=4).value = amount / total_gross if total_gross > 0 else 0
                
                ws3.cell(row=row_num, column=3).number_format = '¥#,##0'
                ws3.cell(row=row_num, column=4).number_format = '0.0%'
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info("Excel export completed successfully")
            return output
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise ValueError(f"Failed to export to Excel: {str(e)}")
    
    @staticmethod
    def _format_currency(amount: float) -> str:
        """Format amount as currency"""
        if amount is None:
            return "¥0"
        return f"¥{amount:,.0f}"
