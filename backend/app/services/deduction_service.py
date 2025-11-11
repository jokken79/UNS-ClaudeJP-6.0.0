"""
Servicio de Deducciones V2.0
============================

Servicio para gestión de deducciones de renta:
- Generar deducciones automáticas mensuales
- Listar deducciones por mes
- Exportar a Excel
- Actualizar estados (pending -> processed -> paid)
- Reportes de cobranza

Autor: Sistema UNS-ClaudeJP
"""

from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func, desc
from typing import List, Optional
from datetime import datetime, date
import io

from app.models.models import (
    User,
    RentDeduction,
    ApartmentAssignment,
    AdditionalCharge,
    Employee,
    Apartment,
    AssignmentStatus,
)
from app.schemas.apartment_v2 import (
    DeductionCreate,
    DeductionResponse,
    DeductionListItem,
    DeductionStatusUpdate,
    DeductionStatus,
)


class DeductionService:
    """Servicio para operaciones de deducciones"""

    def __init__(self, db: Session):
        self.db = db

    # -------------------------------------------------------------------------
    # GESTIÓN DE DEDUCCIONES
    # -------------------------------------------------------------------------

    async def get_monthly_deductions(
        self,
        year: int,
        month: int,
        apartment_id: Optional[int] = None,
        employee_id: Optional[int] = None,
        status: Optional[str] = None,
    ) -> List[DeductionListItem]:
        """
        Obtener deducciones de un mes específico

        Args:
            year: Año
            month: Mes
            apartment_id: Filtrar por apartamento (opcional)
            employee_id: Filtrar por empleado (opcional)
            status: Filtrar por estado (opcional)

        Returns:
            Lista de deducciones del mes

        Raises:
            HTTPException: Si el mes/año no es válido
        """
        from fastapi import HTTPException

        # Validar parámetros
        if month < 1 or month > 12:
            raise HTTPException(status_code=400, detail="Mes debe estar entre 1 y 12")

        if year < 2020 or year > 2030:
            raise HTTPException(status_code=400, detail="Año debe estar entre 2020 y 2030")

        # Consultar rent_deductions para el mes
        query = self.db.query(RentDeduction).filter(
            and_(
                RentDeduction.year == year,
                RentDeduction.month == month,
                RentDeduction.deleted_at.is_(None)
            )
        )

        # Aplicar filtros
        if apartment_id:
            query = query.filter(RentDeduction.apartment_id == apartment_id)
        if employee_id:
            query = query.filter(RentDeduction.employee_id == employee_id)
        if status:
            query = query.filter(RentDeduction.status == status)

        # Ordenar por created_at desc
        deductions = query.order_by(desc(RentDeduction.created_at)).all()

        # Construir respuesta
        results = []
        for deduction in deductions:
            results.append(DeductionListItem(
                id=deduction.id,
                assignment_id=deduction.assignment_id,
                employee_id=deduction.employee_id,
                apartment_id=deduction.apartment_id,
                year=deduction.year,
                month=deduction.month,
                base_rent=deduction.base_rent,
                additional_charges=deduction.additional_charges,
                total_deduction=deduction.total_deduction,
                status=deduction.status,
                processed_date=deduction.processed_date,
                paid_date=deduction.paid_date,
                notes=deduction.notes,
                created_at=deduction.created_at,
                updated_at=deduction.updated_at,
            ))

        return results

    async def generate_monthly_deductions(
        self,
        year: int,
        month: int,
        user_id: int,
    ) -> List[DeductionResponse]:
        """
        Generar deducciones automáticas para un mes

        Proceso:
        1. Buscar asignaciones activas en el mes
        2. Para cada asignación:
           - Calcular días ocupados
           - Calcular renta prorrateada
           - Sumar cargos adicionales approved
           - Crear deducción
        3. Evitar duplicados (UNIQUE constraint)

        Args:
            year: Año
            month: Mes
            user_id: ID del usuario

        Returns:
            Lista de deducciones generadas

        Raises:
            HTTPException: Si ya existen deducciones para el mes
        """
        from fastapi import HTTPException
        import calendar

        # Verificar que no existen deducciones para el mes
        existing = self.db.query(RentDeduction).filter(
            and_(
                RentDeduction.year == year,
                RentDeduction.month == month,
                RentDeduction.deleted_at.is_(None)
            )
        ).first()
        if existing:
            raise HTTPException(
                status_code=409,
                detail=f"Ya existen deducciones generadas para {year}/{month:02d}"
            )

        # Buscar asignaciones activas
        assignments = self.db.query(ApartmentAssignment).filter(
            and_(
                ApartmentAssignment.status == AssignmentStatus.ACTIVE,
                ApartmentAssignment.deleted_at.is_(None)
            )
        ).all()

        created_deductions = []

        # Para cada asignación, crear deducción
        for assignment in assignments:
            # Calcular días del mes
            days_in_month = calendar.monthrange(year, month)[1]

            # Determinar días ocupados
            start_of_month = date(year, month, 1)
            end_of_month = date(year, month, days_in_month)

            # Si la asignación empieza este mes, calcular desde start_date
            if assignment.start_date.year == year and assignment.start_date.month == month:
                days_occupied = (end_of_month - assignment.start_date).days + 1
                base_rent = int((assignment.monthly_rent / days_in_month) * days_occupied)
            else:
                # Mes completo
                days_occupied = days_in_month
                base_rent = assignment.monthly_rent

            # Sumar cargos adicionales aprobados del mes
            charges_sum = self.db.query(func.sum(AdditionalCharge.amount)).filter(
                and_(
                    AdditionalCharge.assignment_id == assignment.id,
                    AdditionalCharge.status == 'approved',
                    AdditionalCharge.charge_date >= start_of_month,
                    AdditionalCharge.charge_date <= end_of_month,
                    AdditionalCharge.deleted_at.is_(None)
                )
            ).scalar() or 0

            # Calcular total
            total_deduction = base_rent + charges_sum

            # Crear deducción
            deduction = RentDeduction(
                assignment_id=assignment.id,
                employee_id=assignment.employee_id,
                apartment_id=assignment.apartment_id,
                year=year,
                month=month,
                base_rent=base_rent,
                additional_charges=charges_sum,
                total_deduction=total_deduction,
                status=DeductionStatus.PENDING,
                created_at=datetime.now(),
            )

            self.db.add(deduction)
            created_deductions.append(deduction)

        # Commit todas las deducciones
        self.db.commit()

        # Refrescar y construir respuesta
        results = []
        for deduction in created_deductions:
            self.db.refresh(deduction)
            results.append(DeductionResponse(
                id=deduction.id,
                assignment_id=deduction.assignment_id,
                employee_id=deduction.employee_id,
                apartment_id=deduction.apartment_id,
                year=deduction.year,
                month=deduction.month,
                base_rent=deduction.base_rent,
                additional_charges=deduction.additional_charges,
                total_deduction=deduction.total_deduction,
                status=deduction.status,
                processed_date=deduction.processed_date,
                paid_date=deduction.paid_date,
                notes=deduction.notes,
                created_at=deduction.created_at,
                updated_at=deduction.updated_at,
            ))

        return results

    async def get_deduction(
        self,
        deduction_id: int,
    ) -> DeductionResponse:
        """
        Obtener deducción por ID

        Args:
            deduction_id: ID de la deducción

        Returns:
            Deducción con detalles

        Raises:
            HTTPException: Si no se encuentra
        """
        from fastapi import HTTPException

        # Consultar rent_deductions
        deduction = self.db.query(RentDeduction).filter(
            and_(
                RentDeduction.id == deduction_id,
                RentDeduction.deleted_at.is_(None)
            )
        ).first()

        if not deduction:
            raise HTTPException(
                status_code=404,
                detail="Deducción no encontrada"
            )

        return DeductionResponse(
            id=deduction.id,
            assignment_id=deduction.assignment_id,
            employee_id=deduction.employee_id,
            apartment_id=deduction.apartment_id,
            year=deduction.year,
            month=deduction.month,
            base_rent=deduction.base_rent,
            additional_charges=deduction.additional_charges,
            total_deduction=deduction.total_deduction,
            status=deduction.status,
            processed_date=deduction.processed_date,
            paid_date=deduction.paid_date,
            notes=deduction.notes,
            created_at=deduction.created_at,
            updated_at=deduction.updated_at,
        )

    async def update_deduction_status(
        self,
        deduction_id: int,
        update: DeductionStatusUpdate,
        current_user: User,
    ) -> DeductionResponse:
        """
        Actualizar estado de deducción

        Estados válidos:
        - pending -> processed
        - processed -> paid
        - paid -> processed (revertir)
        - pending -> cancelled

        Args:
            deduction_id: ID de la deducción
            update: Datos de actualización
            current_user: Usuario actual

        Returns:
            Deducción actualizada

        Raises:
            HTTPException: Si no tiene permisos o validaciones fallidas
        """
        from fastapi import HTTPException

        # Obtener deducción
        deduction = self.db.query(RentDeduction).filter(
            and_(
                RentDeduction.id == deduction_id,
                RentDeduction.deleted_at.is_(None)
            )
        ).first()

        if not deduction:
            raise HTTPException(
                status_code=404,
                detail="Deducción no encontrada"
            )

        # Verificar permisos según el estado
        allowed_roles = ["SUPER_ADMIN", "ADMIN", "COORDINATOR"]
        if current_user.role not in allowed_roles:
            raise HTTPException(
                status_code=403,
                detail="No tiene permisos para actualizar deducciones"
            )

        # Si se marca como paid, solo ADMIN+
        if update.new_status == DeductionStatus.PAID:
            if current_user.role not in ["SUPER_ADMIN", "ADMIN"]:
                raise HTTPException(
                    status_code=403,
                    detail="Solo ADMIN puede marcar deducciones como pagadas"
                )

        # Validar transición de estados
        valid_transitions = {
            DeductionStatus.PENDING: [DeductionStatus.PROCESSED, DeductionStatus.CANCELLED],
            DeductionStatus.PROCESSED: [DeductionStatus.PAID, DeductionStatus.PENDING],
            DeductionStatus.PAID: [DeductionStatus.PROCESSED],
            DeductionStatus.CANCELLED: [],
        }

        if update.new_status not in valid_transitions.get(deduction.status, []):
            raise HTTPException(
                status_code=400,
                detail=f"Transición inválida: {deduction.status} -> {update.new_status}"
            )

        # Actualizar estado
        deduction.status = update.new_status
        deduction.updated_at = datetime.now()

        # Actualizar fechas según el estado
        if update.new_status == DeductionStatus.PROCESSED:
            deduction.processed_date = date.today()
        elif update.new_status == DeductionStatus.PAID:
            deduction.paid_date = date.today()

        # Actualizar notas si se proporcionan
        if update.notes:
            deduction.notes = update.notes

        self.db.commit()
        self.db.refresh(deduction)

        return DeductionResponse(
            id=deduction.id,
            assignment_id=deduction.assignment_id,
            employee_id=deduction.employee_id,
            apartment_id=deduction.apartment_id,
            year=deduction.year,
            month=deduction.month,
            base_rent=deduction.base_rent,
            additional_charges=deduction.additional_charges,
            total_deduction=deduction.total_deduction,
            status=deduction.status,
            processed_date=deduction.processed_date,
            paid_date=deduction.paid_date,
            notes=deduction.notes,
            created_at=deduction.created_at,
            updated_at=deduction.updated_at,
        )

    # -------------------------------------------------------------------------
    # EXPORTACIÓN
    # -------------------------------------------------------------------------

    async def export_deductions_excel(
        self,
        year: int,
        month: int,
        apartment_id: Optional[int],
        user_id: int,
    ) -> bytes:
        """
        Exportar deducciones del mes a Excel

        Args:
            year: Año
            month: Mes
            apartment_id: Filtrar por apartamento (opcional)
            user_id: ID del usuario

        Returns:
            Archivo Excel en bytes

        Raises:
            HTTPException: Si no tiene permisos
        """
        from fastapi import HTTPException
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        # Obtener deducciones
        deductions = await self.get_monthly_deductions(year, month, apartment_id)

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Deducciones {year}-{month:02d}"

        # Escribir headers
        headers = [
            "ID Deducción", "Empleado ID", "Empleado",
            "Apartamento ID", "Apartamento",
            "Año", "Mes", "Renta Base",
            "Cargos Adicionales", "Total Deducción",
            "Estado", "Fecha Procesado", "Fecha Pagado", "Notas"
        ]

        # Estilo de headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Escribir datos
        for row_idx, deduction in enumerate(deductions, 2):
            # Obtener datos relacionados
            employee = self.db.query(Employee).filter(Employee.id == deduction.employee_id).first()
            apartment = self.db.query(Apartment).filter(Apartment.id == deduction.apartment_id).first()

            ws.cell(row=row_idx, column=1, value=deduction.id)
            ws.cell(row=row_idx, column=2, value=deduction.employee_id)
            ws.cell(row=row_idx, column=3, value=employee.full_name_kanji if employee else "N/A")
            ws.cell(row=row_idx, column=4, value=deduction.apartment_id)
            ws.cell(row=row_idx, column=5, value=apartment.name if apartment else "N/A")
            ws.cell(row=row_idx, column=6, value=deduction.year)
            ws.cell(row=row_idx, column=7, value=deduction.month)
            ws.cell(row=row_idx, column=8, value=deduction.base_rent)
            ws.cell(row=row_idx, column=9, value=deduction.additional_charges)
            ws.cell(row=row_idx, column=10, value=deduction.total_deduction)
            ws.cell(row=row_idx, column=11, value=deduction.status)
            ws.cell(row=row_idx, column=12, value=deduction.processed_date.strftime("%Y-%m-%d") if deduction.processed_date else "")
            ws.cell(row=row_idx, column=13, value=deduction.paid_date.strftime("%Y-%m-%d") if deduction.paid_date else "")
            ws.cell(row=row_idx, column=14, value=deduction.notes or "")

        # Auto-ajustar columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Guardar en memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()
