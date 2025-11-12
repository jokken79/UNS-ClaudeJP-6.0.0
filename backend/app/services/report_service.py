"""
Report Service for UNS-ClaudeJP 2.0
Automatic report generation with Excel and charts
"""
import logging
from datetime import datetime
from typing import Dict, List, Optional
from pathlib import Path
from decimal import Decimal

logger = logging.getLogger(__name__)


class ReportService:
    """Servicio de generación automática de reportes en Excel y PDF.

    Genera reportes profesionales con:
    - Excel con formato, fórmulas y gráficos (openpyxl)
    - PDF con diseño profesional (reportlab)
    - Análisis mensuales y anuales de fábrica
    - Recibos de pago individuales

    Attributes:
        reports_dir (Path): Directorio donde se guardan reportes ("reports/")

    Note:
        - Reportes Excel incluyen gráficos automáticos
        - PDF requiere fuentes japonesas para caracteres correctos
        - Todos los reportes se guardan en directorio "reports/"
        - Nombres de archivo incluyen fecha/ID para organización

    Examples:
        >>> service = ReportService()
        >>> result = service.generate_monthly_factory_report(
        ...     factory_id="F001",
        ...     year=2025,
        ...     month=10,
        ...     payrolls=[...],
        ...     factory_config={...}
        ... )
        >>> print(f"Reporte generado: {result['report_path']}")
    """

    def __init__(self):
        """Inicializa el servicio y crea directorio de reportes si no existe."""
        self.reports_dir = Path("reports")
        self.reports_dir.mkdir(exist_ok=True)
    
    def generate_monthly_factory_report(
        self,
        factory_id: str,
        year: int,
        month: int,
        payrolls: List[Dict],
        factory_config: Dict
    ) -> Dict:
        """
        Generate monthly factory report with Excel + charts
        
        Args:
            factory_id: Factory ID
            year: Year
            month: Month
            payrolls: List of payroll data for all employees
            factory_config: Factory configuration
            
        Returns:
            Dict with report path and metrics
        """
        logger.info(f"Generating monthly report for {factory_id} - {year}/{month}")
        
        try:
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Calculate metrics
            total_hours = sum(p['hours']['total_hours'] for p in payrolls)
            total_cost = sum(p['gross_pay'] for p in payrolls)

            actual_revenues = [
                p.get('factory_payment')
                for p in payrolls
                if isinstance(p.get('factory_payment'), (int, float))
            ]

            if actual_revenues:
                total_revenue = sum(actual_revenues)
            else:
                total_revenue = total_hours * factory_config.get('jikyu_tanka', 1500) * 1.2
            
            profit = total_revenue - total_cost
            profit_margin = (profit / total_revenue * 100) if total_revenue > 0 else 0
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "月次レポート"
            
            # Styling
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            subheader_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = f"{factory_config.get('name', factory_id)} - {year}年{month}月 月次レポート"
            title_cell.font = Font(size=16, bold=True, color="1F4788")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            
            # Report date
            ws['A2'] = f"作成日: {datetime.now().strftime('%Y年%m月%d日')}"
            ws['A2'].font = Font(italic=True, size=10)
            ws.merge_cells('A2:F2')
            
            # Summary metrics section
            ws['A4'] = "サマリー"
            ws['A4'].font = Font(size=14, bold=True)
            ws['A4'].fill = subheader_fill
            ws.merge_cells('A4:F4')
            
            metrics = [
                ("総労働時間", f"{total_hours:.1f} 時間"),
                ("総人件費", f"¥{total_cost:,.0f}"),
                ("総売上", f"¥{total_revenue:,.0f}"),
                ("利益", f"¥{profit:,.0f}"),
                ("利益率", f"{profit_margin:.1f}%")
            ]
            
            row = 5
            for metric_name, metric_value in metrics:
                ws[f'A{row}'] = metric_name
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                ws[f'A{row}'].border = border
                
                ws[f'B{row}'] = metric_value
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
                ws[f'B{row}'].border = border
                ws.merge_cells(f'B{row}:F{row}')
                row += 1
            
            # Employee details section
            row += 2
            ws[f'A{row}'] = "社員別詳細"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws[f'A{row}'].fill = subheader_fill
            ws.merge_cells(f'A{row}:F{row}')
            
            row += 1
            # Table headers
            headers = ['社員ID', '社員名', '労働時間', '基本給', '残業代', '総支給額']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Employee data
            row += 1
            start_data_row = row
            for payroll in payrolls:
                ws.cell(row=row, column=1, value=payroll.get('employee_id', 'N/A'))
                ws.cell(row=row, column=2, value=payroll.get('employee_name', 'N/A'))
                ws.cell(row=row, column=3, value=f"{payroll['hours']['total_hours']:.1f}")
                ws.cell(row=row, column=4, value=f"¥{payroll['payments']['base_pay']:,.0f}")
                ws.cell(row=row, column=5, value=f"¥{payroll['payments']['overtime_pay']:,.0f}")
                ws.cell(row=row, column=6, value=f"¥{payroll['gross_pay']:,.0f}")
                
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = border
                    if col > 2:
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
                
                row += 1
            
            # Add chart (if there are employees)
            if payrolls:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "社員別総支給額"
                chart.y_axis.title = '金額 (円)'
                chart.x_axis.title = '社員'
                
                data = Reference(ws, min_col=6, min_row=start_data_row-1, max_row=row-1)
                cats = Reference(ws, min_col=2, min_row=start_data_row, max_row=row-1)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.height = 10
                chart.width = 20
                
                ws.add_chart(chart, f"H{start_data_row}")
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            
            # Save report
            report_filename = f"report_{factory_id}_{year}{month:02d}.xlsx"
            report_path = self.reports_dir / report_filename
            wb.save(report_path)
            
            logger.info(f"Report generated: {report_path}")
            
            return {
                "success": True,
                "report_path": str(report_path),
                "report_filename": report_filename,
                "metrics": {
                    "total_hours": total_hours,
                    "total_cost": total_cost,
                    "total_revenue": total_revenue,
                    "profit": profit,
                    "profit_margin": profit_margin
                },
                "employee_count": len(payrolls)
            }
            
        except Exception as e:
            logger.error(f"Failed to generate report: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    
    def generate_employee_payslip_pdf(self, payroll_data: Dict) -> str:
        """
        Generate individual payslip PDF
        
        Args:
            payroll_data: Payroll data for single employee
            
        Returns:
            str: Path to generated PDF
        """
        logger.info(f"Generating payslip for employee {payroll_data.get('employee_id')}")
        
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.pdfgen import canvas
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # Create PDF
            employee_id = payroll_data.get('employee_id', 'UNKNOWN')
            year = payroll_data.get('year')
            month = payroll_data.get('month')
            
            filename = f"payslip_{employee_id}_{year}{month:02d}.pdf"
            pdf_path = self.reports_dir / filename
            
            c = canvas.Canvas(str(pdf_path), pagesize=A4)
            width, height = A4
            
            # Title
            c.setFont("Helvetica-Bold", 20)
            c.drawCentredString(width/2, height - 40*mm, "給与明細書")
            
            # Date
            c.setFont("Helvetica", 10)
            c.drawString(40*mm, height - 50*mm, f"{year}年{month}月分")
            
            # Employee info
            y = height - 70*mm
            c.setFont("Helvetica-Bold", 12)
            c.drawString(40*mm, y, f"社員ID: {employee_id}")
            y -= 10*mm
            c.drawString(40*mm, y, f"社員名: {payroll_data.get('employee_name', 'N/A')}")
            
            # Hours section
            y -= 20*mm
            c.setFont("Helvetica-Bold", 14)
            c.drawString(40*mm, y, "労働時間")
            
            y -= 8*mm
            c.setFont("Helvetica", 10)
            hours = payroll_data['hours']
            c.drawString(40*mm, y, f"通常時間: {hours['normal_hours']:.1f}h")
            c.drawString(110*mm, y, f"残業時間: {hours['overtime_hours']:.1f}h")
            
            y -= 6*mm
            c.drawString(40*mm, y, f"深夜時間: {hours['night_hours']:.1f}h")
            c.drawString(110*mm, y, f"休日時間: {hours['holiday_hours']:.1f}h")
            
            # Payments section
            y -= 15*mm
            c.setFont("Helvetica-Bold", 14)
            c.drawString(40*mm, y, "支給")
            
            y -= 8*mm
            c.setFont("Helvetica", 10)
            payments = payroll_data['payments']
            c.drawString(40*mm, y, f"基本給: ¥{payments['base_pay']:,.0f}")
            
            y -= 6*mm
            c.drawString(40*mm, y, f"残業手当: ¥{payments['overtime_pay']:,.0f}")
            
            y -= 6*mm
            c.drawString(40*mm, y, f"深夜手当: ¥{payments['night_pay']:,.0f}")
            
            y -= 6*mm
            c.drawString(40*mm, y, f"休日手当: ¥{payments['holiday_pay']:,.0f}")
            
            # Bonuses
            if payroll_data.get('bonuses', {}).get('total', 0) > 0:
                y -= 6*mm
                c.drawString(40*mm, y, f"手当合計: ¥{payroll_data['bonuses']['total']:,.0f}")
            
            # Deductions section
            y -= 15*mm
            c.setFont("Helvetica-Bold", 14)
            c.drawString(40*mm, y, "控除")
            
            y -= 8*mm
            c.setFont("Helvetica", 10)
            deductions = payroll_data['deductions']
            c.drawString(40*mm, y, f"社会保険: ¥{deductions.get('insurance', 0):,.0f}")
            
            y -= 6*mm
            c.drawString(40*mm, y, f"所得税: ¥{deductions.get('tax', 0):,.0f}")
            
            y -= 6*mm
            c.drawString(40*mm, y, f"寮費: ¥{deductions.get('apartment', 0):,.0f}")
            
            # Total
            y -= 15*mm
            c.setFont("Helvetica-Bold", 16)
            c.drawString(40*mm, y, f"差引支給額: ¥{payroll_data['net_pay']:,.0f}")
            
            # Footer
            c.setFont("Helvetica", 8)
            c.drawString(40*mm, 30*mm, "この給与明細書は自動生成されています。")
            c.drawString(40*mm, 25*mm, f"発行日: {datetime.now().strftime('%Y年%m月%d日')}")
            
            c.save()
            
            logger.info(f"Payslip PDF generated: {pdf_path}")
            return str(pdf_path)
            
        except Exception as e:
            logger.error(f"Failed to generate payslip PDF: {e}")
            return ""
    
    def generate_annual_summary_report(
        self,
        factory_id: str,
        year: int,
        monthly_data: List[Dict]
    ) -> Dict:
        """
        Generate annual summary report
        
        Args:
            factory_id: Factory ID
            year: Year
            monthly_data: List of monthly metrics
            
        Returns:
            Dict with report info
        """
        logger.info(f"Generating annual report for {factory_id} - {year}")
        
        try:
            from openpyxl import Workbook
            from openpyxl.chart import LineChart, Reference
            
            wb = Workbook()
            ws = wb.active
            ws.title = "年次サマリー"
            
            # Title
            ws.merge_cells('A1:E1')
            ws['A1'] = f"{factory_id} - {year}年 年次レポート"
            ws['A1'].font = Font(size=16, bold=True)
            
            # Headers
            headers = ['月', '労働時間', '人件費', '売上', '利益']
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=3, column=col_idx, value=header)
            
            # Monthly data
            row = 4
            for month_data in monthly_data:
                ws.cell(row=row, column=1, value=f"{month_data['month']}月")
                ws.cell(row=row, column=2, value=month_data['total_hours'])
                ws.cell(row=row, column=3, value=month_data['total_cost'])
                ws.cell(row=row, column=4, value=month_data['total_revenue'])
                ws.cell(row=row, column=5, value=month_data['profit'])
                row += 1
            
            # Add trend chart
            chart = LineChart()
            chart.title = "年間推移"
            chart.y_axis.title = "金額 (円)"
            chart.x_axis.title = "月"
            
            data = Reference(ws, min_col=3, min_row=3, max_row=row-1, max_col=5)
            cats = Reference(ws, min_col=1, min_row=4, max_row=row-1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "G3")
            
            # Save
            report_filename = f"annual_report_{factory_id}_{year}.xlsx"
            report_path = self.reports_dir / report_filename
            wb.save(report_path)
            
            logger.info(f"Annual report generated: {report_path}")
            
            return {
                "success": True,
                "report_path": str(report_path),
                "report_filename": report_filename
            }
            
        except Exception as e:
            logger.error(f"Failed to generate annual report: {e}")
            return {
                "success": False,
                "error": str(e)
            }


    # =========================================================================
    # APARTMENT MODULE REPORTS
    # =========================================================================

    def get_occupancy_report(self, prefecture: Optional[str] = None,
                            building_name: Optional[str] = None) -> dict:
        """
        Reporte de ocupación de apartamentos

        Args:
            prefecture: Prefectura a filtrar (opcional)
            building_name: Nombre del edificio a filtrar (opcional)

        Returns:
            Dict con métricas de ocupación
        """
        from sqlalchemy.orm import Session
        from sqlalchemy import func, and_
        from app.core.database import SessionLocal
        from app.models.models import Apartment, ApartmentAssignment, AssignmentStatus

        db = SessionLocal()

        try:
            # Query base de apartamentos
            query = db.query(Apartment).filter(Apartment.deleted_at.is_(None))

            # Aplicar filtros si se proporcionan
            if prefecture:
                query = query.filter(Apartment.prefecture == prefecture)

            if building_name:
                query = query.filter(Apartment.building_name == building_name)

            # Métricas principales
            total_apartments = query.count()

            # Apartamentos con asignaciones activas (ocupados)
            occupied_query = query.join(ApartmentAssignment).filter(
                and_(
                    ApartmentAssignment.status == AssignmentStatus.ACTIVE,
                    ApartmentAssignment.deleted_at.is_(None)
                )
            )
            occupied_apartments = occupied_query.distinct(Apartment.id).count()

            # Apartamentos vacantes
            vacant_apartments = total_apartments - occupied_apartments

            # Tasa de ocupación
            occupancy_rate = (occupied_apartments / total_apartments * 100) if total_apartments > 0 else 0.0

            # Capacidad total vs ocupación actual (por simplicidad, 1 empleado por apartamento)
            total_capacity = total_apartments
            current_occupancy = occupied_apartments

            # Desglose por prefectura
            by_prefecture = db.query(
                Apartment.prefecture,
                func.count(Apartment.id).label('total'),
                func.count(ApartmentAssignment.id).label('occupied')
            ).outerjoin(
                ApartmentAssignment,
                and_(
                    Apartment.id == ApartmentAssignment.apartment_id,
                    ApartmentAssignment.status == AssignmentStatus.ACTIVE,
                    ApartmentAssignment.deleted_at.is_(None)
                )
            ).filter(
                Apartment.deleted_at.is_(None)
            ).group_by(Apartment.prefecture).all()

            # Desglose por tipo de habitación
            by_room_type = db.query(
                Apartment.room_type,
                func.count(Apartment.id).label('total'),
                func.count(ApartmentAssignment.id).label('occupied')
            ).outerjoin(
                ApartmentAssignment,
                and_(
                    Apartment.id == ApartmentAssignment.apartment_id,
                    ApartmentAssignment.status == AssignmentStatus.ACTIVE,
                    ApartmentAssignment.deleted_at.is_(None)
                )
            ).filter(
                Apartment.deleted_at.is_(None)
            ).group_by(Apartment.room_type).all()

            return {
                "total_apartments": total_apartments,
                "occupied_apartments": occupied_apartments,
                "vacant_apartments": vacant_apartments,
                "occupancy_rate": round(occupancy_rate, 2),
                "total_capacity": total_capacity,
                "current_occupancy": current_occupancy,
                "by_prefecture": [
                    {
                        "prefecture": p[0] or "N/A",
                        "total": p[1],
                        "occupied": p[2] or 0,
                        "occupancy_rate": round((p[2] or 0) / p[1] * 100, 2) if p[1] > 0 else 0.0
                    }
                    for p in by_prefecture
                ],
                "by_room_type": [
                    {
                        "room_type": r[0] or "N/A",
                        "total": r[1],
                        "occupied": r[2] or 0,
                        "occupancy_rate": round((r[2] or 0) / r[1] * 100, 2) if r[1] > 0 else 0.0
                    }
                    for r in by_room_type
                ]
            }
        finally:
            db.close()

    def get_arrears_report(self, year: int, month: int) -> dict:
        """
        Reporte de pagos pendientes (morosidad)

        Args:
            year: Año
            month: Mes

        Returns:
            Dict con métricas de cobranza
        """
        from sqlalchemy.orm import Session
        from sqlalchemy import func, and_
        from app.core.database import SessionLocal
        from app.models.models import RentDeduction, Employee, DeductionStatus

        db = SessionLocal()

        try:
            # Deducciones del mes
            deductions = db.query(RentDeduction).filter(
                and_(
                    RentDeduction.year == year,
                    RentDeduction.month == month,
                    RentDeduction.deleted_at.is_(None)
                )
            ).all()

            # Métricas principales
            total_to_collect = sum([d.total_deduction for d in deductions])

            paid_deductions = [d for d in deductions if d.status == DeductionStatus.PAID]
            total_paid = sum([d.total_deduction for d in paid_deductions])

            pending_deductions = [d for d in deductions if d.status in [DeductionStatus.PENDING, DeductionStatus.PROCESSED]]
            total_pending = sum([d.total_deduction for d in pending_deductions])

            employees_with_debt = len(set([d.employee_id for d in pending_deductions]))

            # Top deudores
            top_debtors_data = db.query(
                RentDeduction.employee_id,
                Employee.full_name_kanji,
                func.sum(RentDeduction.total_deduction).label('total_debt')
            ).join(
                Employee,
                RentDeduction.employee_id == Employee.id
            ).filter(
                and_(
                    RentDeduction.year == year,
                    RentDeduction.month == month,
                    RentDeduction.status.in_([DeductionStatus.PENDING, DeductionStatus.PROCESSED]),
                    RentDeduction.deleted_at.is_(None)
                )
            ).group_by(
                RentDeduction.employee_id,
                Employee.full_name_kanji
            ).order_by(
                func.sum(RentDeduction.total_deduction).desc()
            ).limit(10).all()

            top_debtors = [
                {
                    "employee_id": d[0],
                    "employee_name": d[1],
                    "total_debt": d[2]
                }
                for d in top_debtors_data
            ]

            return {
                "total_to_collect": total_to_collect,
                "total_paid": total_paid,
                "total_pending": total_pending,
                "employees_with_debt": employees_with_debt,
                "collection_rate": round((total_paid / total_to_collect * 100) if total_to_collect > 0 else 0.0, 2),
                "top_debtors": top_debtors
            }
        finally:
            db.close()

    def get_maintenance_report(self) -> dict:
        """
        Reporte de mantenimiento y cargos adicionales

        Returns:
            Dict con métricas de mantenimiento
        """
        from sqlalchemy.orm import Session
        from sqlalchemy import func, and_, extract
        from app.core.database import SessionLocal
        from app.models.models import AdditionalCharge, Apartment
        from datetime import datetime, timedelta

        db = SessionLocal()

        try:
            # Todos los cargos adicionales
            all_charges = db.query(AdditionalCharge).filter(
                AdditionalCharge.deleted_at.is_(None)
            ).all()

            total_charges = len(all_charges)

            # Desglose por tipo de cargo
            by_charge_type = db.query(
                AdditionalCharge.charge_type,
                func.count(AdditionalCharge.id).label('count'),
                func.sum(AdditionalCharge.amount).label('total_amount')
            ).filter(
                AdditionalCharge.deleted_at.is_(None)
            ).group_by(AdditionalCharge.charge_type).all()

            # Tendencias mensuales (últimos 6 meses)
            six_months_ago = datetime.now() - timedelta(days=180)
            monthly_trends = db.query(
                extract('year', AdditionalCharge.charge_date).label('year'),
                extract('month', AdditionalCharge.charge_date).label('month'),
                func.count(AdditionalCharge.id).label('count'),
                func.sum(AdditionalCharge.amount).label('total_amount')
            ).filter(
                and_(
                    AdditionalCharge.charge_date >= six_months_ago.date(),
                    AdditionalCharge.deleted_at.is_(None)
                )
            ).group_by(
                extract('year', AdditionalCharge.charge_date),
                extract('month', AdditionalCharge.charge_date)
            ).order_by(
                extract('year', AdditionalCharge.charge_date),
                extract('month', AdditionalCharge.charge_date)
            ).all()

            # Apartamentos con más incidentes
            apartments_with_most_incidents = db.query(
                AdditionalCharge.apartment_id,
                Apartment.name,
                func.count(AdditionalCharge.id).label('incident_count'),
                func.sum(AdditionalCharge.amount).label('total_cost')
            ).join(
                Apartment,
                AdditionalCharge.apartment_id == Apartment.id
            ).filter(
                AdditionalCharge.deleted_at.is_(None)
            ).group_by(
                AdditionalCharge.apartment_id,
                Apartment.name
            ).order_by(
                func.count(AdditionalCharge.id).desc()
            ).limit(10).all()

            return {
                "total_charges": total_charges,
                "by_charge_type": {
                    ct[0]: {
                        "count": ct[1],
                        "total_amount": ct[2] or 0
                    }
                    for ct in by_charge_type
                },
                "monthly_trends": [
                    {
                        "year": int(mt[0]),
                        "month": int(mt[1]),
                        "count": mt[2],
                        "total_amount": mt[3] or 0
                    }
                    for mt in monthly_trends
                ],
                "apartments_with_most_incidents": [
                    {
                        "apartment_id": a[0],
                        "apartment_name": a[1],
                        "incident_count": a[2],
                        "total_cost": a[3] or 0
                    }
                    for a in apartments_with_most_incidents
                ]
            }
        finally:
            db.close()

    def get_cost_analysis_report(self, year: int, month: int) -> dict:
        """
        Análisis de costos (solo ADMIN)

        Args:
            year: Año
            month: Mes

        Returns:
            Dict con análisis de costos
        """
        from sqlalchemy.orm import Session
        from sqlalchemy import func, and_
        from app.core.database import SessionLocal
        from app.models.models import Apartment, RentDeduction

        db = SessionLocal()

        try:
            # Costos totales (rentas de todos los apartamentos)
            all_apartments = db.query(Apartment).filter(
                Apartment.deleted_at.is_(None)
            ).all()

            total_costs = sum([
                (apt.base_rent or 0) + (apt.management_fee or 0)
                for apt in all_apartments
            ])

            # Deducciones totales (lo que cobramos)
            deductions = db.query(RentDeduction).filter(
                and_(
                    RentDeduction.year == year,
                    RentDeduction.month == month,
                    RentDeduction.deleted_at.is_(None)
                )
            ).all()

            total_deductions = sum([d.total_deduction for d in deductions])

            # Margen de beneficio (simplificado)
            profit_margin = ((total_deductions - total_costs) / total_costs * 100) if total_costs > 0 else 0.0

            # Promedio por apartamento
            average_per_apartment = total_deductions / len(deductions) if deductions else 0

            # Tendencias de costos (últimos 6 meses)
            cost_trends = []
            for i in range(6, 0, -1):
                target_month = month - i
                target_year = year
                if target_month <= 0:
                    target_month += 12
                    target_year -= 1

                month_deductions = db.query(
                    func.sum(RentDeduction.total_deduction)
                ).filter(
                    and_(
                        RentDeduction.year == target_year,
                        RentDeduction.month == target_month,
                        RentDeduction.deleted_at.is_(None)
                    )
                ).scalar() or 0

                cost_trends.append({
                    "year": target_year,
                    "month": target_month,
                    "total_deductions": month_deductions
                })

            return {
                "total_costs": total_costs,
                "total_deductions": total_deductions,
                "profit_margin": round(profit_margin, 2),
                "average_per_apartment": round(average_per_apartment, 2),
                "cost_trends": cost_trends
            }
        finally:
            db.close()


# Global instance
report_service = ReportService()
