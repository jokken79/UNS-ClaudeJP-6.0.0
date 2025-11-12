"""Import/Export API Endpoints for UNS-ClaudeJP 2.0."""
import logging
import shutil
import tempfile
from io import BytesIO
from pathlib import Path
from typing import Final, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.import_service import import_service
from app.services.auth_service import AuthService
from app.core.config import settings
from app.core.database import get_db

router = APIRouter()
logger = logging.getLogger(__name__)

UPLOAD_DIR: Final[Path] = Path(settings.UPLOAD_DIR) / "import_temp"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

STRICT_SECURITY_ENVS = {"production", "staging"}

if settings.ENVIRONMENT.lower() in STRICT_SECURITY_ENVS and not settings.DEBUG:
    _admin_guard_dependency = AuthService.require_role("admin")
else:

    async def _admin_guard_dependency():  # pragma: no cover - trivial helper
        """Allow unauthenticated imports in non-production environments."""

        return None


def _write_upload_to_temp(upload: UploadFile, expected_suffixes: tuple[str, ...]) -> Path:
    """Persist an uploaded file to a temporary location with a safe filename."""

    original_name = upload.filename or ""
    suffix = Path(original_name).suffix.lower()
    if suffix not in expected_suffixes:
        raise HTTPException(
            status_code=400,
            detail=f"Only files with extensions {', '.join(expected_suffixes)} are supported",
        )

    try:
        temp_file = tempfile.NamedTemporaryFile(
            delete=False, suffix=suffix, dir=UPLOAD_DIR
        )
        with temp_file as buffer:
            upload.file.seek(0)
            shutil.copyfileobj(upload.file, buffer)
        return Path(temp_file.name)
    except HTTPException:
        raise
    except Exception:  # pragma: no cover - defensive programming
        logger.exception("Failed to persist uploaded file")
        raise HTTPException(status_code=500, detail="Unable to process uploaded file")


def _create_template_response(
    columns: list[str],
    *,
    sheet_name: str,
    filename: str,
    sample_rows: Optional[list[list[str]]] = None,
) -> StreamingResponse:
    """Generate an in-memory Excel workbook with the provided columns."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for column_index, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = max(len(column_title) + 4, 16)

    if sample_rows:
        for row_offset, row_values in enumerate(sample_rows, start=2):
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_offset, column=column_index, value=value)

    worksheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-cache",
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/employees")
async def import_employees(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(_admin_guard_dependency)
):
    """
    Import employees from Excel file

    Expected columns: 派遣元ID, 氏名, フリガナ, 生年月日, 性別, 国籍, etc.

    Requires admin role.
    """
    temp_file: Optional[Path] = None

    try:
        temp_file = _write_upload_to_temp(file, (".xlsx", ".xls"))

        logger.info(f"Importing employees from {file.filename}")

        # Import
        results = import_service.import_employees_from_excel(str(temp_file), db)

        return results
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importing employees: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if temp_file and temp_file.exists():
            try:
                temp_file.unlink()
            except OSError as e:
                # Silently ignore cleanup errors (file may be locked or already deleted)
                logger.debug(f"Could not delete temp file {temp_file}: {e}")
                pass


@router.post("/timer-cards")
async def import_timer_cards(
    file: UploadFile = File(...),
    factory_id: str = Query(...),
    year: int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db),
    current_user = Depends(_admin_guard_dependency)
):
    """
    Import timer cards from Excel file

    Expected columns: 日付, 社員ID, 社員名, 出勤時刻, 退勤時刻

    Requires admin role.
    """
    temp_file: Optional[Path] = None

    try:
        temp_file = _write_upload_to_temp(file, (".xlsx", ".xls"))

        logger.info(f"Importing timer cards for {factory_id} - {year}/{month}")
        
        results = import_service.import_timer_cards_from_excel(
            str(temp_file),
            factory_id,
            year,
            month,
            db,
        )
        
        return results
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importing timer cards: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if temp_file and temp_file.exists():
            try:
                temp_file.unlink()
            except OSError as e:
                # Silently ignore cleanup errors (file may be locked or already deleted)
                logger.debug(f"Could not delete temp file {temp_file}: {e}")
                pass


@router.post("/factory-configs")
async def import_factory_configs(
    directory_path: str,
    db: Session = Depends(get_db),
    current_user=Depends(_admin_guard_dependency),
):
    """
    Import factory configurations from JSON files
    
    Args:
        directory_path: Path to directory containing factory JSON files
    """
    try:
        results = import_service.import_factory_configs_from_json(directory_path, db)
        return results
        
    except Exception as e:
        logger.error(f"Error importing factory configs: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/template/employees")
async def download_employee_template():
    """Download Excel template for employee import"""
    try:
        columns = [
            "派遣元ID",
            "氏名",
            "フリガナ",
            "ローマ字",
            "生年月日 (YYYY-MM-DD)",
            "性別",
            "国籍",
            "郵便番号",
            "住所",
            "携帯電話",
            "電話番号",
            "メール",
            "在留カード番号",
            "ビザ種類",
            "ビザ期限 (YYYY-MM-DD)",
        ]
        sample_rows = [[
            "FAC-001",
            "山田 太郎",
            "ヤマダ タロウ",
            "Yamada Taro",
            "1990-01-01",
            "男性",
            "日本",
            "123-4567",
            "東京都港区1-2-3",
            "090-1234-5678",
            "03-1234-5678",
            "taro.yamada@example.com",
            "AB1234567",
            "就労",
            "2026-12-31",
        ]]

        return _create_template_response(
            columns,
            sheet_name="Employees",
            filename="empleados_template.xlsx",
            sample_rows=sample_rows,
        )

    except Exception as exc:  # pragma: no cover - defensive
        logger.exception("Error creating employee template")
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/template/timer-cards")
async def download_timer_cards_template():
    """Download Excel template for timer cards import"""
    try:
        columns = [
            "日付 (YYYY-MM-DD)",
            "社員ID",
            "社員名",
            "出勤時刻 (HH:MM)",
            "退勤時刻 (HH:MM)",
        ]
        sample_rows = [[
            "2024-01-01",
            "EMP-001",
            "山田 太郎",
            "09:00",
            "18:00",
        ]]

        return _create_template_response(
            columns,
            sheet_name="TimerCards",
            filename="tarjetas_tiempo_template.xlsx",
            sample_rows=sample_rows,
        )

    except Exception as exc:  # pragma: no cover - defensive
        logger.exception("Error creating timer cards template")
        raise HTTPException(status_code=500, detail=str(exc))
